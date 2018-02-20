"""
Reading in the excel data for the chimps spreadsheet to
create the proper proportions
"""

from openpyxl import load_workbook

wb = load_workbook('Chimps.xlsx', data_only = True)

intervals = wb.worksheets[1]
proportions = wb.worksheets[0]

start_col = input('Enter the starting column: ')
end_col = start_col + 24
start_row = input('Enter the starting row: ')
end_row = start_row + 23

for col in range(start_col, end_col):
	for row in range(start_row, end_row):
		interactions = 0
		has_numbers = False
		i = row
		maximum = i + 276
		while i <= maximum:
			cell = intervals.cell(row = i, column = col)
			value = cell.value
			if value == 0:
				has_numbers = True
			elif value == 1:
				has_numbers = True
				interactions += 1

			i += 23

		if has_numbers == False:
			proportions.cell(row = row, column = (col + 7), value = 'XX')
		else:
			proportions.cell(row = row, column = (col + 7), value = interactions)


wb.save('Chimps3.xlsx')

